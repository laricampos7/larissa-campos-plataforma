#!/usr/bin/env python3
# Gera o app de cada aluno a partir da planilha modelo-treinos.xlsx
# Uso:  python gerar_treinos.py
import json, re, os, unicodedata, sys
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE, "..", "index.html")     # o app serve de modelo
XLSX = os.path.join(BASE, "modelo-treinos.xlsx")
OUT = os.path.join(BASE, "alunos")

DESC_PADRAO = ("Você está na fase de maior intensidade do seu ciclo. Foco em cargas "
               "progressivas, técnica impecável e recuperação de qualidade. Estamos "
               "construindo o pico do seu desempenho.")

# textinho menor (Sora) dentro dos cards de stat (Volume / Frequência)
SMALL = '<small style="font-family:Sora;font-size:13px;color:var(--txt-dim)">%s</small>'

def stat_inner(raw, num_suffix):
    """Conteúdo do <b> de um card de stat a partir do valor da planilha.
    - vazio       -> None (mantém o padrão do template)
    - só número   -> "16" + num_suffix  (ex.: '16 séries/grupo', '5x / semana')
    - número+texto-> número em destaque + resto menor ('2x na semana + beach tênis')
    - texto livre -> mostrado como está
    """
    raw = (raw or "").strip()
    if not raw:
        return None
    if re.fullmatch(r"[\d.,]+", raw):
        return raw + (SMALL % num_suffix)
    m = re.match(r"^([\d.,]+\s*x?)\s+(.*)$", raw, re.I)
    if m:
        return m.group(1).replace(" ", "") + (SMALL % (" " + m.group(2).strip()))
    return raw

def nums_in(s):
    """Todos os números de um texto. '7-8' -> [7,8]; '3x12' -> [3,12]; vazio -> []."""
    return [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", str(s or ""))]

def slug(s):
    s = unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s or "aluno"

def linhas(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    head = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({head[i]: ("" if r[i] is None else str(r[i]).strip()) for i in range(len(head))})
    return out

# ---------- Avaliação física (aba "Avaliacao") ----------
def num(s):
    """'78,1' -> 78.1 ; vazio/invalido -> None"""
    s = str(s).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def fmt(v, dec=1):
    """78.1 -> '78,1' ; 29.0 -> '29' (inteiro sem casas)"""
    if v is None:
        return ""
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return ("%.*f" % (dec, v)).replace(".", ",")

GOOD_STYLE = "background:color-mix(in srgb,var(--pos) 18%,transparent);color:var(--pos)"
DIM_STYLE = "color:var(--txt-faint)"

def _delta(r):
    """retorna (texto_badge, melhorou:bool) ou (None, None) se sem dados de início/atual"""
    ini, at = num(r.get("inicio", "")), num(r.get("atual", ""))
    if ini is None or at is None:
        return None, None
    unidade = (r.get("unidade", "") or "").strip()
    direcao = ((r.get("direcao", "") or "baixo")).strip().lower()
    d = at - ini
    arrow = "▲" if d > 0 else ("▼" if d < 0 else "•")
    melhorou = (d < 0 and direcao == "baixo") or (d > 0 and direcao == "alto")
    unid = "pts" if unidade == "%" else unidade
    txt = arrow + " " + fmt(abs(d)) + ((" " + unid) if unid else "")
    return txt, melhorou

def meas_row(r):
    label = (r.get("metrica", "") or "").strip()
    unidade = (r.get("unidade", "") or "").strip()
    meta = (r.get("meta", "") or "").strip()
    val = fmt(num(r.get("atual", ""))) + ((" " + unidade) if unidade else "")
    lbl = label
    if meta:
        lbl += ('<i style="color:var(--txt-faint);font-style:normal;font-size:11px">'
                ' · meta ' + meta + ((" " + unidade) if unidade else "") + '</i>')
    txt, melhorou = _delta(r)
    badge = ""
    if txt:
        style = GOOD_STYLE if melhorou else DIM_STYLE
        badge = '<small class="up" style="' + style + '">' + txt + '</small>'
    return ('<div class="meas-row"><span class="l">' + lbl + '</span>'
            '<div class="v"><b>' + val + '</b>' + badge + '</div></div>')

def kpi_card(titulo, r, rodape):
    unidade = (r.get("unidade", "") or "").strip()
    txt, melhorou = _delta(r)
    delta = ""
    if txt:
        cls = "delta up" if melhorou else "delta"
        st = "" if melhorou else ' style="color:var(--txt-dim)"'
        delta = '<span class="' + cls + '"' + st + '>' + txt + '</span>'
    return ('<div class="card glow kpi"><h3>' + titulo.upper() + '</h3>'
            '<b>' + fmt(num(r.get("atual", ""))) +
            '<small style="font-family:Sora;font-size:14px"> ' + unidade + '</small></b>'
            + delta + '<small>' + rodape + '</small></div>')

def aplicar_avaliacao(html, rows):
    """Injeta composição corporal + circunferências (e remove o gráfico de exemplo)."""
    comp = [r for r in rows if "compos" in (r.get("secao", "") or "").lower()]
    med = [r for r in rows if r not in comp]

    if comp:
        inner = "<h3>COMPOSIÇÃO CORPORAL · INÍCIO → AGORA</h3>" + "".join(meas_row(r) for r in comp)
        bloco = '<div class="card glow" id="evalComp">' + inner + '<!--/evalComp--></div>'
        html = re.sub(r'<div class="card glow" id="evalComp">.*?<!--/evalComp--></div>',
                      lambda m: bloco, html, count=1, flags=re.S)
        # KPIs do topo (Início) a partir da composição
        idx = {(r.get("metrica", "") or "").strip().lower(): r for r in comp}
        defs = [("Peso atual", "peso", "desde o início"),
                ("% Gordura", "% gordura", "avaliação"),
                ("Massa magra", "massa magra", "músculo"),
                ("Gordura (kg)", "gordura (kg)", "massa gorda")]
        cards = [kpi_card(t, idx[k], rod) for (t, k, rod) in defs if k in idx]
        if len(cards) >= 3:
            kpi = ('<div class="grid g4" id="kpiTop" style="margin-bottom:16px">'
                   + "".join(cards) + '<!--/kpiTop--></div>')
            html = re.sub(r'<div class="grid g4" id="kpiTop".*?<!--/kpiTop--></div>',
                          lambda m: kpi, html, count=1, flags=re.S)

    if med:
        inner = "<h3>MEDIDAS (cm) · INÍCIO → AGORA</h3>" + "".join(meas_row(r) for r in med)
        bloco = '<div class="card glow" id="evalMed">' + inner + '<!--/evalMed--></div>'
        html = re.sub(r'<div class="card glow" id="evalMed">.*?<!--/evalMed--></div>',
                      lambda m: bloco, html, count=1, flags=re.S)

    # remove o gráfico de exemplo (dados fictícios de 12 semanas)
    html = re.sub(r'\s*<div class="card glow" id="evalChart".*?<!--/evalChart--></div>',
                  "", html, count=1, flags=re.S)
    return html


def avaliacao_placeholder(html):
    """Aluna SEM linhas na aba Avaliacao: remove os blocos de EXEMPLO
    (KPIs do topo, gráfico de 12 semanas, composição e medidas) para não
    exibir dados que não são dela, e mostra um aviso amigável."""
    aviso = ('<div class="card glow" id="evalChart" style="margin-bottom:16px">'
             '<h3>COMPOSIÇÃO CORPORAL</h3>'
             '<p style="color:var(--txt-dim);font-size:14px;line-height:1.6;margin-top:8px">'
             '📋 Sua avaliação física ainda não foi cadastrada. Assim que fizermos a sua '
             'primeira avaliação, os dados de composição corporal e medidas aparecem aqui, '
             'com a sua evolução início → agora.</p>'
             '<!--/evalChart--></div>')
    # troca o gráfico de exemplo pelo aviso
    html = re.sub(r'<div class="card glow" id="evalChart".*?<!--/evalChart--></div>',
                  lambda m: aviso, html, count=1, flags=re.S)
    # remove os KPIs de composição de exemplo do topo
    html = re.sub(r'\s*<div class="grid g4" id="kpiTop".*?<!--/kpiTop--></div>',
                  "", html, count=1, flags=re.S)
    # remove composição e medidas de exemplo
    html = re.sub(r'<div class="card glow" id="evalComp">.*?<!--/evalComp--></div>',
                  "", html, count=1, flags=re.S)
    html = re.sub(r'<div class="card glow" id="evalMed">.*?<!--/evalMed--></div>',
                  "", html, count=1, flags=re.S)
    return html


def main():
    if not os.path.exists(XLSX):
        sys.exit("Planilha não encontrada: " + XLSX)
    if not os.path.exists(TEMPLATE):
        sys.exit("Modelo index.html não encontrado em " + TEMPLATE)

    wb = load_workbook(XLSX, data_only=True)
    alunos = linhas(wb["Alunos"]) if "Alunos" in wb.sheetnames else []
    treinos = linhas(wb["Treinos"]) if "Treinos" in wb.sheetnames else []
    biblio = {}
    if "Biblioteca" in wb.sheetnames:
        for b in linhas(wb["Biblioteca"]):
            biblio[b.get("exercicio", "").lower()] = b
    avaliacoes = {}
    if "Avaliacao" in wb.sheetnames:
        for a in linhas(wb["Avaliacao"]):
            nome_a = (a.get("aluno", "") or "").strip().lower()
            if nome_a:
                avaliacoes.setdefault(nome_a, []).append(a)

    template = open(TEMPLATE, encoding="utf-8").read()
    os.makedirs(OUT, exist_ok=True)
    gerados = 0

    for al in alunos:
        nome = al.get("nome", "").strip()
        if not nome:
            continue
        # monta os dias preservando a ordem das linhas
        dias, ordem = {}, []
        for t in treinos:
            if t.get("aluno", "").strip().lower() != nome.lower():
                continue
            d = t.get("dia", "").strip() or "A"
            if d not in dias:
                dias[d] = {"dia": d, "titulo": t.get("titulo_dia", "").strip() or ("Dia " + d),
                           "label": t.get("titulo_dia", "").strip() or ("Dia " + d), "exercicios": []}
                ordem.append(d)
            ref = biblio.get(t.get("exercicio", "").lower(), {})
            dias[d]["exercicios"].append({
                "nome": t.get("exercicio", "").strip(),
                "grupo": t.get("grupo", "").strip() or ref.get("grupo", ""),
                "series": t.get("series", "").strip(),
                "reps": t.get("reps", "").strip(),
                "rpe": t.get("rpe", "").strip(),
                "dica": t.get("dica", "").strip() or ref.get("dica", ""),
                "video": t.get("video", "").strip() or ref.get("video", ""),
            })
        treino = [dias[d] for d in ordem]
        if not treino:
            print("  (sem treinos para %s — pulando)" % nome)
            continue

        html = template
        # namespace de armazenamento por aluna (isola metas/cargas/fotos/conquistas no navegador)
        html = html.replace("var STORE_NS = 'demo';", "var STORE_NS = '%s';" % slug(nome))
        # aderência: se a coluna "aderencia" estiver preenchida com um valor real, fixa esse valor;
        # vazia ou "—" = o app calcula sozinho pelos treinos marcados nos últimos 30 dias (updateAderencia).
        ader_manual = (al.get("aderencia") or "").strip()
        if ader_manual in ("—", "-", "–"):
            ader_manual = ""
        html = html.replace("var ADER_MANUAL = '';", "var ADER_MANUAL = '%s';" % ader_manual)
        html = html.replace("João Silva", nome)
        # iniciais do avatar
        ini = "".join(w[0] for w in nome.split()[:2]).upper() or "AL"
        html = html.replace('<div class="avatar">JS</div>', '<div class="avatar">%s</div>' % ini)
        if al.get("plano"):     html = html.replace("Plano Premium · 12ª semana", al["plano"])
        if al.get("bloco"):     html = html.replace("Mesociclo atual · Bloco 3 de 4", al["bloco"])
        if al.get("fase"):      html = html.replace("<h2>FASE DE FORÇA</h2>", "<h2>%s</h2>" % al["fase"])
        if al.get("descricao"): html = html.replace(DESC_PADRAO, al["descricao"])
        # ----- cards de stat da aba Ciclo (Volume / Frequência / Intensidade) -----
        # Calcula a partir do próprio treino: total de séries na semana e RPE médio.
        total_series, rpes = 0.0, []
        for d in treino:
            for ex in d["exercicios"]:
                ns = nums_in(ex["series"])
                if ns:
                    total_series += ns[0]            # 1º número = nº de séries (ignora "3x12")
                nr = nums_in(ex["rpe"])
                if nr:
                    rpes.append(sum(nr) / len(nr))    # "7-8" vira 7,5
        auto_vol = int(round(total_series)) if total_series > 0 else None
        auto_rpe = (sum(rpes) / len(rpes)) if rpes else None

        # Frequência: coluna "frequencia"; se vazia, usa o nº de dias de treino do aluno.
        freq_html = stat_inner(al.get("frequencia", ""), "x / semana") or (str(len(treino)) + (SMALL % "x / semana"))
        html = html.replace(
            '<h3>FREQUÊNCIA</h3><b style="font-family:Anton;font-size:28px">5<small style="font-family:Sora;font-size:13px;color:var(--txt-dim)">x / semana</small></b>',
            '<h3>FREQUÊNCIA</h3><b style="font-family:Anton;font-size:28px">%s</b>' % freq_html, 1)
        # Volume: coluna "volume" (manual) tem prioridade; senão, total de séries da semana.
        vol_html = stat_inner(al.get("volume", ""), " séries/semana")
        if not vol_html and auto_vol:
            vol_html = str(auto_vol) + (SMALL % " séries/semana")
        if vol_html:
            html = html.replace(
                '<h3>VOLUME SEMANAL</h3><b style="font-family:Anton;font-size:28px">16<small style="font-family:Sora;font-size:13px;color:var(--txt-dim)"> séries/grupo</small></b>',
                '<h3>VOLUME SEMANAL</h3><b style="font-family:Anton;font-size:28px">%s</b>' % vol_html, 1)
        # Intensidade: coluna "intensidade" (manual) tem prioridade; senão, RPE médio; senão "—".
        intens = (al.get("intensidade", "") or "").strip()
        if intens:
            if re.match(r"^[\d.,]", intens):
                intens = "RPE " + intens
            intens_html = intens
        elif auto_rpe is not None:
            intens_html = "RPE " + fmt(auto_rpe)
        else:
            intens_html = "—"
        html = html.replace(
            '<h3>INTENSIDADE MÉDIA</h3><b style="font-family:Anton;font-size:28px">RPE 8,5</b>',
            '<h3>INTENSIDADE MÉDIA</h3><b style="font-family:Anton;font-size:28px">%s</b>' % intens_html, 1)
        # avaliação física da aluna (aba "Avaliacao")
        aval = avaliacoes.get(nome.lower(), [])
        gordura_atual = ""
        for r in aval:
            if (r.get("metrica", "") or "").strip().lower() == "% gordura":
                gv = num(r.get("atual", ""))
                if gv is not None:
                    gordura_atual = fmt(gv)
        # números do topo (check-in / aderência / treinos / % gordura)
        if any(al.get(k) for k in ("proximo_checkin", "aderencia", "treinos_mes")) or gordura_atual:
            hm = ('<div class="hero-meta">'
                  '<div><span>Próximo check-in</span><b>%s</b></div>'
                  '<div><span>Aderência (30d)</span><b style="color:var(--pos)" id="heroAder">%s</b></div>'
                  '<div><span>Treinos no mês</span><b>%s</b></div>'
                  '<div><span>%% Gordura</span><b>%s</b></div>'
                  '</div>') % (al.get("proximo_checkin") or "—", al.get("aderencia") or "—",
                               al.get("treinos_mes") or "—",
                               (gordura_atual + "<small>%</small>") if gordura_atual else "—")
            html = re.sub(r'<div class="hero-meta">.*?</div>\s*</section>',
                          lambda m: hm + "\n  </section>", html, count=1, flags=re.S)
        # metas do aluno (seção "Metas") — coluna "metas" na aba Alunos.
        # Formato: uma meta por item, separadas por ";" ou quebra de linha.
        # Progresso opcional após "=" (0 a 100). Ex: "Agachamento 20 kg=30; Ganhar 500 g=0"
        metas_raw = al.get("metas", "").strip()
        if metas_raw:
            metas = []
            for it in re.split(r"[;\n]+", metas_raw):
                it = it.strip()
                if not it:
                    continue
                if "=" in it:
                    nome_m, pct = it.rsplit("=", 1)
                    try:
                        pct = int(float(pct.strip()))
                    except ValueError:
                        nome_m, pct = it, 0
                    metas.append({"name": nome_m.strip(), "pct": max(0, min(100, pct))})
                else:
                    metas.append({"name": it, "pct": 0})
            seed_js = "const METAS_SEED = " + json.dumps(metas, ensure_ascii=False) + ";"
            html = re.sub(r"const METAS_SEED = \[.*?\];", lambda m: seed_js, html, count=1, flags=re.S)

        # avaliação física → composição corporal + circunferências (e remove o gráfico de exemplo)
        if aval:
            html = aplicar_avaliacao(html, aval)
        else:
            html = avaliacao_placeholder(html)

        novo_treino = "var TREINO = " + json.dumps(treino, ensure_ascii=False, separators=(",", ":")) + ";"
        html = re.sub(r"^var TREINO = .*;$", lambda m: novo_treino, html, count=1, flags=re.M)

        pasta = os.path.join(OUT, slug(nome))
        os.makedirs(pasta, exist_ok=True)
        with open(os.path.join(pasta, "index.html"), "w", encoding="utf-8") as f:
            f.write(html)
        gerados += 1
        print("  ✓ %s  ->  alunos/%s/index.html  (%d dias, %d exercícios)" %
              (nome, slug(nome), len(treino), sum(len(d["exercicios"]) for d in treino)))

    print("\nPronto! %d app(s) gerado(s) na pasta 'alunos/'." % gerados)
    print("Envie a pasta de cada aluno (ou só o index.html) pra ele abrir no celular.")

if __name__ == "__main__":
    main()
